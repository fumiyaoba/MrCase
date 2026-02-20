from pathlib import Path
from typing import Iterable, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from MRCASE import env
from MRCASE.models.manhour_entry import ManHourEntry

HEADERS = env.EXCEL_HEADERS


def open_sheet(excel_filepath: str | Path, excel_sheet_name: str | None) -> Tuple[Workbook, Worksheet]:
    """
    excel_filepath のブックを開いて、指定シートを返す。
    シート名が無い/見つからない場合は env.EXCEL_SHEET_RAW -> active の順で選ぶ。
    """
    workbook = load_workbook(excel_filepath)

    sheet = None
    if excel_sheet_name and excel_sheet_name in workbook.sheetnames:
        sheet = workbook[excel_sheet_name]
    elif getattr(env, "EXCEL_SHEET_RAW", None) and env.EXCEL_SHEET_RAW in workbook.sheetnames:
        sheet = workbook[env.EXCEL_SHEET_RAW]
    else:
        sheet = workbook.active

    return workbook, sheet


def _find_next_row(worksheet: Worksheet) -> int:
    """
    次に書き込む行番号（1始まり）を返す。
    最後の非空行の次の行。
    """
    max_row = worksheet.max_row

    # シートが完全に空（max_row=1 で A1 も None など）にも耐える
    last_row = 0
    for row_num in range(max_row, 0, -1):
        row = worksheet[row_num]
        if any(cell.value is not None for cell in row):
            last_row = row_num
            break

    return last_row + 1


def _ensure_header(worksheet: Worksheet) -> None:
    """
    1行目が空ならヘッダを書き込む。
    """
    first_row_values = [worksheet.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
    if all(v is None for v in first_row_values):
        worksheet.append(HEADERS)


def append_entries(
    entries: Iterable[ManHourEntry],
    excel_filepath: str | Path,
    excel_sheet_name: str | None = None,
) -> None:
    """
    entries を末尾に追記して保存する。
    """
    workbook, ws = open_sheet(excel_filepath, excel_sheet_name)

    _ensure_header(ws)

    next_row = _find_next_row(ws)

    for e in entries:
        text_raw = getattr(e, "text_raw", "")  # 無ければ空
        ws.cell(row=next_row, column=1, value=e.work_date)
        ws.cell(row=next_row, column=2, value=e.project)
        ws.cell(row=next_row, column=3, value=e.assignee)
        ws.cell(row=next_row, column=4, value=float(e.hours))
        ws.cell(row=next_row, column=5, value=text_raw)
        next_row += 1

    workbook.save(excel_filepath)